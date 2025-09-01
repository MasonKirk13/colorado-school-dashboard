import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import re

def normalize(col):
    return re.sub(r'[^a-z0-9]', '', str(col).lower())

wb = openpyxl.load_workbook('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - Suppressed - with combined years.xlsx')
ref_ws = wb['2024-2025 Data']
header_row = [cell.value for cell in next(ref_ws.iter_rows(min_row=1, max_row=1))]
header_norm = [normalize(h) for h in header_row]

# List of year sheets to process
year_sheets = [
    '2019-2020 Data',
    '2018-2019 Data',
    '2017-2018 Data',
    '2016-2017 Data',
    '2015-2016 Data',
    '2014-2015 Data',
]

for sheetname in year_sheets:
    ws = wb[sheetname]
    data = list(ws.values)
    if not data:
        continue
    columns = [str(c) for c in data[0]]
    df = pd.DataFrame(data[1:], columns=columns)
    # Build a mapping from normalized col name to actual col name
    df_norm_map = {normalize(c): c for c in df.columns}
    # Reorder and filter columns to match reference header
    out = pd.DataFrame()
    for h, hn in zip(header_row, header_norm):
        if hn in df_norm_map:
            out[h] = df[df_norm_map[hn]]
        else:
            out[h] = np.nan
    # Remove all columns not in the reference header
    ws_new = wb.create_sheet(title=sheetname + ' (fixed2)')
    ws_new.append(header_row)
    for row in dataframe_to_rows(out, index=False, header=False):
        ws_new.append(row)
    print(f'Processed and fixed: {sheetname}')

wb.save('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL2.xlsx')
print('Saved as 2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL2.xlsx')
