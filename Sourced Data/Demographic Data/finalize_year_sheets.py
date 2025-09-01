import openpyxl

wb = openpyxl.load_workbook('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL.xlsx')

# List of year sheets
year_sheets = [
    '2019-2020 Data',
    '2018-2019 Data',
    '2017-2018 Data',
    '2016-2017 Data',
    '2015-2016 Data',
    '2014-2015 Data',
]

# Delete original year sheets
for sheet in year_sheets:
    if sheet in wb.sheetnames:
        del wb[sheet]
# Rename (fixed) sheets back to original names
for sheet in year_sheets:
    fixed = sheet + ' (fixed)'
    if fixed in wb.sheetnames:
        ws = wb[fixed]
        ws.title = sheet

wb.save('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL.xlsx')
print('Deleted old year sheets and renamed fixed sheets to original names.')
