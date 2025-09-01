import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# Open the workbook
wb = openpyxl.load_workbook('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - Suppressed - with combined years.xlsx')

# Get the reference header and style from the 2024-2025 Data sheet
ref_ws = wb['2024-2025 Data']
header_row = [cell.value for cell in next(ref_ws.iter_rows(min_row=1, max_row=1))]
header_style = [cell._style for cell in next(ref_ws.iter_rows(min_row=1, max_row=1))]

# List of new year sheets to process
year_sheets = [
    '2019-2020 Data',
    '2018-2019 Data',
    '2017-2018 Data',
    '2016-2017 Data',
    '2015-2016 Data',
    '2014-2015 Data',
]

for sheetname in year_sheets:
    ws = wb[sheetname]
    # Read data into DataFrame
    data = list(ws.values)
    if not data:
        continue
    columns = [str(c) for c in data[0]]
    df = pd.DataFrame(data[1:], columns=columns)
    # Reorder and filter columns to match reference header
    out = pd.DataFrame()
    for h in header_row:
        if h in df.columns:
            out[h] = df[h]
        else:
            out[h] = np.nan
    # Remove all columns not in the reference header
    ws_new = wb.create_sheet(title=sheetname + ' (fixed)')
    # Write header with style
    ws_new.append(header_row)
    for i, cell in enumerate(ws_new[1]):
        cell._style = header_style[i]
    # Write data
    for row in dataframe_to_rows(out, index=False, header=False):
        ws_new.append(row)
    print(f'Processed and fixed: {sheetname}')

wb.save('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL.xlsx')
print('Saved as 2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL.xlsx')
