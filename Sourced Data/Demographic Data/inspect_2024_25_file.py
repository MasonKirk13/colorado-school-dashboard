import openpyxl

wb = openpyxl.load_workbook('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - Suppressed.xlsx', read_only=True, data_only=True)
print('Sheet names:', wb.sheetnames)
ws = wb.active
print('Active sheet title:', ws.title)
print('First 5 rows:')
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(row)
    if i >= 4:
        break
