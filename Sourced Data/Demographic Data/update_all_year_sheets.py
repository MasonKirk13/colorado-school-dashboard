import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

wb = openpyxl.load_workbook('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL2.xlsx')
ref_ws = wb['2024-2025 Data']
header_row = [cell.value for cell in next(ref_ws.iter_rows(min_row=1, max_row=1))]

# List of year sheets to process
year_sheets = [
    '2019-2020 Data',
    '2018-2019 Data',
    '2017-2018 Data',
    '2016-2017 Data',
    '2015-2016 Data',
    '2014-2015 Data',
]

def transform(df):
    out = pd.DataFrame()
    out['Organization Code'] = df.get('district code', df.get('Organization Code'))
    out['Organization Name'] = df.get('district name', df.get('Organization Name'))
    out['School Code'] = df.get('school code')
    out['School Name'] = df.get('school name')
    out['Lowest Grade Level'] = df.get('lowest grade', df.get('Lowest Grade Level'))
    out['Highest Grade Level'] = df.get('highest grade', df.get('Highest Grade Level'))
    out['Charter Y/N'] = 'N/A'
    out['PK-12 Total'] = df.get('pk-12 total')
    out['Free Lunch'] = df.get('free lunch')
    out['Reduced Lunch'] = df.get('reduced lunch')
    out['Paid Lunch'] = df.get('not eligible')
    # Demographic sums
    out['Amer Indian/Alaskan Native'] = df.get('american indian or alaskan native female', 0).fillna(0) + df.get('american indian or alaskan native male', 0).fillna(0)
    out['Asian'] = df.get('asian female', 0).fillna(0) + df.get('asian male', 0).fillna(0)
    out['Black or African American'] = df.get('black or african american female', 0).fillna(0) + df.get('black or african american male', 0).fillna(0)
    out['Hispanic or Latino'] = df.get('hispanic or latino female', 0).fillna(0) + df.get('hispanic or latino male', 0).fillna(0)
    out['White'] = df.get('white female', 0).fillna(0) + df.get('white male', 0).fillna(0)
    out['Hawaiian/Pacific Islander'] = df.get('native hawaiian or other pacific islander female', 0).fillna(0) + df.get('native hawaiian or other pacific islander male', 0).fillna(0)
    out['Two or More Races'] = df.get('two or more races female', 0).fillna(0) + df.get('two or more races male', 0).fillna(0)
    out['Female'] = (
        df.get('american indian or alaskan native female', 0).fillna(0) +
        df.get('asian female', 0).fillna(0) +
        df.get('black or african american female', 0).fillna(0) +
        df.get('hispanic or latino female', 0).fillna(0) +
        df.get('white female', 0).fillna(0) +
        df.get('native hawaiian or other pacific islander female', 0).fillna(0) +
        df.get('two or more races female', 0).fillna(0)
    )
    out['Male'] = (
        df.get('american indian or alaskan native male', 0).fillna(0) +
        df.get('asian male', 0).fillna(0) +
        df.get('black or african american male', 0).fillna(0) +
        df.get('hispanic or latino male', 0).fillna(0) +
        df.get('white male', 0).fillna(0) +
        df.get('native hawaiian or other pacific islander male', 0).fillna(0) +
        df.get('two or more races male', 0).fillna(0)
    )
    out['Non-Binary'] = 0
    # Ensure column order
    out = out[header_row]
    return out

for sheetname in year_sheets:
    ws = wb[sheetname]
    data = list(ws.values)
    columns = [str(c) for c in data[0]]
    df = pd.DataFrame(data[1:], columns=columns)
    out = transform(df)
    # Remove old sheet and add new one
    wb.remove(ws)
    ws_new = wb.create_sheet(title=sheetname)
    ws_new.append(header_row)
    for row in dataframe_to_rows(out, index=False, header=False):
        ws_new.append(row)
    print(f'Updated: {sheetname}')

wb.save('2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL3.xlsx')
print('Saved as 2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - FINAL3.xlsx')
