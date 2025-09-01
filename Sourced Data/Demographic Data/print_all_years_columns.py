import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

main_path = '2024-25_FRL_Race_Gender_bySchoolandSchoolFlags - Suppressed.xlsx'
years = [
    ('2014_15_PK_12_FreeReducedLunchEligibilitybyDistrictandSchool_2.xlsx', '2014_15_sch_membershipbysch_race_ethnicity_gender_grade.xlsx', '2014-15'),
    ('2015_16_PK_12_FreeReducedLunchEligibilitybySchool_1.xlsx', '2015_16_sch_membershipbysch_race_ethnicity_gender_grade_1.xlsx', '2015-16'),
    ('2016-17_PK-12_PupilMembership_bySchool_FRL_0.xlsx', '2016-17_sch_membershipbysch_race_ethnicity_gender_grade.xlsx', '2016-17'),
    ('2017-18_PK12_FRL_bySchool_0.xlsx', '2017-18-Membership-Race-Gender-byGradeSchool.xlsx', '2017-18'),
    ('2018-19_PK12_FRL_bySchool.xlsx', '2018-19-Membership-Race-Gender-byGradeSchool.xlsx', '2018-19'),
    ('2019-20_PK12_FRL_bySchool2a.xlsx', '2019-20-Membership-Race-Gender-byGradeSchool.xlsx', '2019-20'),
]

# Get reference headers from 2024-25 sheet
ref_headers = pd.read_excel(main_path, sheet_name=0, nrows=0).columns.tolist()

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def add_header_borders(ws, num_cols):
    for col in range(1, num_cols+1):
        ws.cell(row=1, column=col).border = border

def normalize_columns(df):
    return {str(c).strip().upper().replace('\t','').replace('  ',' '): c for c in df.columns}

def merge_and_write():
    for frl_file, mem_file, sheet_name in years:
        if not os.path.exists(frl_file) or not os.path.exists(mem_file):
            print(f"Missing file for {sheet_name}")
            continue
        frl = pd.read_excel(frl_file, header=2)
        mem = pd.read_excel(mem_file, header=2)
        frl_map = normalize_columns(frl)
        mem_map = normalize_columns(mem)
        # Find best merge keys
        org_code_frl = frl_map.get('DISTRICT CODE') or frl_map.get('ORG CODE') or frl_map.get('ORG. CODE')
        org_code_mem = mem_map.get('ORG CODE') or mem_map.get('DISTRICT CODE') or mem_map.get('ORG. CODE')
        school_code_frl = frl_map.get('SCHOOL CODE')
        school_code_mem = mem_map.get('SCHOOL CODE')
        # Rename for merge
        if org_code_frl and org_code_mem:
            frl = frl.rename(columns={org_code_frl: 'Org. Code'})
            mem = mem.rename(columns={org_code_mem: 'Org. Code'})
        if school_code_frl and school_code_mem:
            frl = frl.rename(columns={school_code_frl: 'School Code'})
            mem = mem.rename(columns={school_code_mem: 'School Code'})
        # Convert merge keys to string
        for k in ['Org. Code', 'School Code']:
            if k in frl.columns:
                frl[k] = frl[k].astype(str)
            if k in mem.columns:
                mem[k] = mem[k].astype(str)
        # Merge
        if 'Org. Code' in frl.columns and 'Org. Code' in mem.columns and 'School Code' in frl.columns and 'School Code' in mem.columns:
            merged = pd.merge(mem, frl, on=['Org. Code', 'School Code'], how='outer', suffixes=('_mem', '_frl'))
        else:
            print(f"Could not find merge keys for {sheet_name}, using membership data only.")
            merged = mem
        out = pd.DataFrame(columns=ref_headers)
        for col in ref_headers:
            if col in merged.columns:
                out[col] = merged[col]
            else:
                out[col] = ''
        with pd.ExcelWriter(main_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            out.to_excel(writer, sheet_name=sheet_name, index=False)
        wb = load_workbook(main_path)
        ws = wb[sheet_name]
        add_header_borders(ws, len(ref_headers))
        wb.save(main_path)
        print(f"Added merged and formatted sheet for {sheet_name}")
    print("All year sheets merged, formatted, and added.")

# Delete all sheets except the original
wb = load_workbook(main_path)
original_sheets = ['2024-2025 Data']
for sheet in wb.sheetnames:
    if sheet not in original_sheets:
        del wb[sheet]
wb.save(main_path)
print(f"Deleted all sheets except: {original_sheets}")

if __name__ == "__main__":
    merge_and_write()
